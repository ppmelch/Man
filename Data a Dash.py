import dash
import dash_bootstrap_components as dbc
from dash import Input, Output, dcc, html, dash_table
import openpyxl
import pandas as pd
from dash.dependencies import Input, Output

# Load data from Excel
excel_dataframe = openpyxl.load_workbook("NewData.xlsx")
dataframe = excel_dataframe.active

# Convert data from Excel to pandas DataFrame
data = []
for row in range(1, dataframe.max_row + 1):
    _row = [dataframe.cell(row=row, column=1).value]
    for col in range(2, dataframe.max_column + 1):
        _row.append(dataframe.cell(row=row, column=col).value)
    data.append(_row)

headers = ["State"] + ["2011", "2012", "2013", "2014", "2015", "2016", "2017", "2018", "2019", "2020", "2021", "2022", "2023"]
df = pd.DataFrame(data, columns=headers)

app = dash.Dash(external_stylesheets=[dbc.themes.BOOTSTRAP])
SIDEBAR_STYLE = {
    "position": "fixed",
    "top": 0,
    "left": 0,
    "bottom": 0,
    "bottom":"#f8f9fa",
    "width": "16rem",
    "padding": "2rem 1rem",
    "background-color": "#f8f9fa",  # Cambia el color de fondo de la barra lateral
    "border-right": "2px solid #b4001f",  # Cambia el color del borde de la barra lateral
}

CONTENT_STYLE = {
    "margin-left": "18rem",
    "margin-right": "2rem",
    "padding": "2rem 1rem",
    "background-color": "#ecf0f1",  # Cambia el color de fondo del contenido principal
}

sidebar = html.Div(
    [
        html.H2("Security CRIME INCIDENCE ", className="display-4", style={"color": "#b4001f", "font-size": "40px"}),  # Cambia el color y tamaño de letra del título
        html.Hr(),
        html.P("Nacional Avarage", className="display-4",style={"color": "#030303", "font-size": "25px"}),
        dbc.Nav(
            [
                dbc.NavLink("Home", href="/", active="exact",style={"color": "#b4001f"}),
                dbc.NavLink("Dashboard", href="/dashboard", active="exact",style={"color": "#b4001f"}),
                dbc.NavLink("Excel", href="/table", active="exact" , style={"color": "#b4001f"}),
                dbc.NavLink("Map", href="/Map", active="exact", style={"color": "#b4001f"}),  # Cambia el color del texto del botón en la barra de navegación
            ],
            vertical=True,
            pills=True,
        ),
    ],
    style=SIDEBAR_STYLE,
)


content = html.Div(id="page-content", style=CONTENT_STYLE)

app.layout = html.Div([dcc.Location(id="url"), sidebar, content])

# Move the callback inside the layout for the dashboard
@app.callback(
    [Output('line-chart', 'figure'),
     Output('data-table', 'columns'),
     Output('data-table', 'data')],
    [Input('dropdown-state', 'value'),
     Input('dropdown-year', 'value')]
)
def update_chart(selected_states, selected_years):
    # Set default values for selected_states and selected_years if they are None
    selected_states = selected_states or []
    selected_years = selected_years or df.columns[1:].tolist()  # Use all available years if none are selected

    # Filter out None values in the selected_states list
    selected_states = [state for state in selected_states if state is not None]

    if selected_states:
        # Filter DataFrame by selected states
        filtered_df = df[df['State'].isin(selected_states)]

        # Create line chart
        traces = [
            {
                'x': filtered_df.columns[1:],
                'y': filtered_df.loc[filtered_df['State'] == state].iloc[0, 1:],
                'type': 'line',
                'name': state
            } for state in selected_states
        ]

        figure = {'data': traces, 'layout': {'title': 'Datos por Estado'}}

        # Define columns for DataTable based on selected years
        table_columns = [{'name': 'State', 'id': 'State'}] + [{'name': year, 'id': year} for year in selected_years]

        # Filter data for DataTable based on selected years
        table_data = [{'State': state} for state in selected_states]
        for year in selected_years:
            for state in selected_states:
                if year in filtered_df.columns:
                    table_data[selected_states.index(state)][year] = filtered_df.loc[filtered_df['State'] == state, year].values[0]
                else:
                    table_data[selected_states.index(state)][year] = None

        return figure, table_columns, table_data
    else:
        return {'data': [], 'layout': {}}, [], []


@app.callback(Output("page-content", "children"), [Input("url", "pathname")])
def render_page_content(pathname):
    if pathname == "/":
        return html.P("")
    elif pathname == "/dashboard":
        return [
            #Grafica TEXTO 
            html.H1(("CRIME RATE (PREELIMINARY SELECTED LOCATIONS)"),style={"color":"#b4001f"}),
            html.P(("2011 to 2023"),style={"color":"#474747"}),
            html.P(("Rate per 100k inhabitants"),style={"color":"#5b5b5b"}),
            dcc.Dropdown(
                id='dropdown-state',
                options=[
                    {'label': state, 'value': state} for state in df['State']
                ],
                multi=True,
                placeholder='Seleccione estado(s)'
            ),
            dcc.Dropdown(
                id='dropdown-year',
                options=[
                    {'label': year, 'value': year} for year in df.columns[1:]
                ],
                multi=True,  # Allow multi-selection
                placeholder='Seleccione uno o más años'
            ),
            dcc.Graph(
                id='line-chart',
            ),
            dash_table.DataTable(
                id='data-table',
                style_table={'overflowX': 'auto', 'width': '30%', 'float': 'left'},
            )
        ]
    elif pathname == "/table":
        return [
            html.H1(("Security Crimes Table"),style={"color":"#b4001f"}),
            html.P(("")),
            html.P(("")),
            dash_table.DataTable(id='data-table', data=df.to_dict('records'), columns=[{'name': col, 'id': col} for col in df.columns])
        ]
    elif pathname == "/Map":
        return html.H1(("Color Code Map"),style={"color":"#b4001f"})
    else:
        return html.Div(
            [
                html.H1("404: Not found", className="text-danger"),
                html.Hr(),
                html.P(f"The pathname {pathname} was not recognized..."),
            ],
            className="p-3 bg-light rounded-3",
        )

if __name__ == "__main__":
    app.run_server(port=8888)





